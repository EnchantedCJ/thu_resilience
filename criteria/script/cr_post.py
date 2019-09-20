# -*- coding: utf-8 -*-
from openpyxl import load_workbook

__all__ = ['postprocess']


def postprocess(resultDir, excelFile):
    wb = load_workbook(excelFile)

    # building
    workDir = resultDir + '/building'
    for pga, sheet in zip([0.2, 0.3, 0.4], ['Sheet1', 'Sheet2', 'Sheet3']):
        ws = wb[sheet]
        levels = []
        downtimes = []

        resultFile = workDir + '/' + str(pga) + '/blg2criteria_' + str(pga) + '.txt'
        with open(resultFile, 'r', encoding='utf-8') as f:
            for line in f.readlines():
                temp = line.strip('\n').split()
                levels.append(float(temp[0]))
                downtimes.append(float(temp[1]))

        for i in range(8):
            row = str(i + 3)
            ws['F' + row] = levels[i]
            ws['G' + row] = downtimes[i]

    # transport
    workDir = resultDir + '/transport'
    resultwb = load_workbook(workDir + '/output.xlsx')
    for rrow, sheet in zip(['1', '2', '3'], ['Sheet1', 'Sheet2', 'Sheet3']):
        ws = wb[sheet]
        levels = []
        downtimes = []

        columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for i in range(8):
            col = columns[i + 1]
            levels.append(resultwb['system_dur'][col + rrow].value)
            col = columns[i]
            downtimes.append(resultwb['restore'][col + rrow].value)

        for i in range(8):
            row = str(i + 11)
            ws['F' + row] = levels[i]
            ws['G' + row] = downtimes[i]

    # lifeline
    workDir = resultDir + '/lifeline'
    for lifeline, startline in zip(['供电', '供水', '供暖'], [19, 27, 43]):
        for pga, sheet in zip([0.2, 0.3, 0.4], ['Sheet1', 'Sheet2', 'Sheet3']):
            resultwb = load_workbook(workDir + '/输出数据/' + str(pga) + 'g条件下' + lifeline + '系统支撑的建筑系统功能恢复时间.xlsx')
            ws = wb[sheet]
            levels = []
            downtimes = []

            for i in range(8):
                row = str(i + 1)
                levels.append(resultwb['Sheet1']['B' + row].value)
                downtimes.append(resultwb['Sheet1']['C' + row].value)

            for i in range(8):
                row = str(i + startline)
                ws['F' + row] = levels[i]
                ws['G' + row] = downtimes[i]

    wb.save(excelFile)


if __name__ == '__main__':
    resultDir = './results'
    excelFile = './平台整体指标文件.xlsx'
    postprocess(resultDir, excelFile)
